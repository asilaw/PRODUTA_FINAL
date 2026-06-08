"""
modules/financial_calc.py
Kalkulasi kelayakan finansial investasi kapasitas filling line.

Sumber data:
  - Komponen mesin & harga     : data/Financial_Param.xlsx (MACHINE COMPONENT sheet)
  - Overhead CAPEX breakdown   : data/Financial_Param.xlsx (CAPEX GENERAL sheet)
  - OPEX manpower              : data/Financial_Param.xlsx (OPEX GENERAL sheet)
  - Parameter finansial default: data/Financial_Param.xlsx (FINANCIAL sheet)
  - Referensi CAPEX total      : Pak Ardi FBMI ~Rp 11.9B (Juni 2026)
  - Maintenance                : Pak Ardi FBMI Rp 20M/bulan = Rp 240M/tahun
  - Model FCF                  : Fonterra/Lactalis Group Valuation Model (IRR_RainWaterHarvesting)
    FCF = EBIT + Dep - CashTax - MaintCapex
        = (Benefit - OPEX - Dep) + Dep - CashTax - MaintCapex
        = Benefit - OPEX - CashTax - MaintCapex
"""
import numpy as np
import numpy_financial as npf
from pathlib import Path
import json

# ── Rupiah formatter ──────────────────────────────────────────────────────────
def fmt_rp(amount: float) -> str:
    amount = round(amount)
    return "Rp " + f"{amount:,}".replace(",", ".")


# ── Load dari Financial_Param.xlsx jika tersedia, fallback ke hardcode ────────
_FP_PATH = Path("data/Financial_Param.xlsx")

def _load_fp_machines() -> dict:
    """Load MACHINE COMPONENT sheet dari Financial_Param.xlsx."""
    try:
        import openpyxl
        wb  = openpyxl.load_workbook(_FP_PATH, read_only=True, data_only=True)
        ws  = wb["MACHINE COMPONENT"]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip().upper() if h else "" for h in rows[0]]
        machines = {}
        for row in rows[1:]:
            if not row[0]:
                continue
            d = dict(zip(header, row))
            # key unik dari nama (lowercase, spasi→underscore, max 30 char)
            key = str(d.get("MACHINE NAME","")).lower()[:30].replace(" ","_").replace("/","_").replace("-","_")
            key = ''.join(c for c in key if c.isalnum() or c == '_')
            if not key:
                continue
            fmt_compat = str(d.get("FORMAT COMPATIBILITY","")).strip().strip("|")
            line_compat = str(d.get("EXISTING LINE COMPATIBILITY","")).strip().strip("|")
            machines[key] = {
                "name":             str(d.get("COMPONENT","")).strip(),
                "full_name":        str(d.get("MACHINE NAME","")).strip(),
                "capex":            int(float(d.get("BASE CAPEX",0) or 0)),
                "opex_per_ton":     float(d.get("BASE OPEX/TON",0) or 0),
                "capacity_impact":  float(d.get("CAPACITY IMPACT (%)",0) or 0),
                "capacity_kg_hr":   float(d.get("CAPACITY (kg/hr)",0) or 0),
                "format_compat":    [f.strip() for f in fmt_compat.split("|") if f.strip()],
                "line_compat":      [l.strip() for l in line_compat.split("|") if l.strip()],
                "is_core":          str(d.get("IS_CORE","NO")).strip().upper() == "YES",
                "opex_rate":        0.05,   # fallback maintenance rate
                "role":             str(d.get("COMPONENT","")).strip(),
                "url":              "",
            }
        wb.close()
        return machines
    except Exception:
        return {}

def _load_fp_capex_general() -> dict:
    """Load CAPEX GENERAL sheet."""
    try:
        import openpyxl
        wb  = openpyxl.load_workbook(_FP_PATH, read_only=True, data_only=True)
        ws  = wb["CAPEX GENERAL"]
        result = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[0] != "CAPEX_TYPE":
                result[str(row[0]).strip()] = {
                    "basis": str(row[1]).strip() if row[1] else "",
                    "value": float(row[2]) if row[2] is not None else 0.0,
                }
        wb.close()
        return result
    except Exception:
        return {}

def _load_fp_opex_general() -> dict:
    """Load OPEX GENERAL sheet."""
    try:
        import openpyxl
        wb  = openpyxl.load_workbook(_FP_PATH, read_only=True, data_only=True)
        ws  = wb["OPEX GENERAL"]
        result = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[0] != "OPEX_TYPE":
                result[str(row[0]).strip()] = {
                    "basis": str(row[1]).strip() if row[1] else "",
                    "value": float(row[2]) if row[2] is not None else 0.0,
                }
        wb.close()
        return result
    except Exception:
        return {}

def _load_fp_financial() -> dict:
    """Load FINANCIAL sheet."""
    try:
        import openpyxl
        wb  = openpyxl.load_workbook(_FP_PATH, read_only=True, data_only=True)
        ws  = wb["FINANCIAL"]
        result = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[0] != "PARAMETER":
                result[str(row[0]).strip()] = row[1]
        wb.close()
        return result
    except Exception:
        return {}


# ── Load data dari xlsx ───────────────────────────────────────────────────────
_fp_machines  = _load_fp_machines()
_fp_capex_gen = _load_fp_capex_general()
_fp_opex_gen  = _load_fp_opex_general()
_fp_financial = _load_fp_financial()

# ── MACHINES: pakai dari xlsx jika berhasil, fallback ke hardcode ─────────────
if _fp_machines:
    MACHINES = _fp_machines
else:
    # Fallback hardcode (data Pak Ardi Juni 2026)
    MACHINES = {
        "shiputec_multilane": {
            "name": "FILLER PLATFORM", "full_name": "SHIPUTEC MULTI-LANE POWDER SACHET",
            "capex": 729_600_000, "opex_rate": 0.09, "opex_per_ton": 260_000,
            "capacity_kg_hr": 350, "is_core": True,
            "format_compat": ["SSS"], "line_compat": ["D"], "role": "Filling", "url": "",
        },
        "shiputec_stickpack": {
            "name": "FILLER PLATFORM", "full_name": "SHIPUTEC SPMP-480 MULTILANE STICKPACK",
            "capex": 620_800_000, "opex_rate": 0.09, "opex_per_ton": 220_000,
            "capacity_kg_hr": 120, "is_core": True,
            "format_compat": ["STICKPACK"], "line_compat": ["STICKPACK"], "role": "Filling", "url": "",
        },
        "wolf_vpc250": {
            "name": "FILLER PLATFORM", "full_name": "WOLF VPC-250 VERTICAL PACKAGING MACHINE",
            "capex": 1_410_400_000, "opex_rate": 0.09, "opex_per_ton": 180_000,
            "capacity_kg_hr": 220, "is_core": True,
            "format_compat": ["SSS","BIB"], "line_compat": ["B","G"], "role": "Filling", "url": "",
        },
        "checkweigher": {
            "name": "INSPECTION SYSTEM 1", "full_name": "INLINE CHECKWEIGHER WITH AIR REJECTOR",
            "capex": 48_000_000, "opex_rate": 0.06, "opex_per_ton": 55_000,
            "capacity_kg_hr": 300, "is_core": True,
            "format_compat": ["SSS","BIB","STICKPACK"], "line_compat": ["B","G","STICKPACK"], "role": "Inspeksi", "url": "",
        },
        "xray": {
            "name": "INSPECTION SYSTEM 2", "full_name": "XRAY FOREIGN BODY DETECTION SYSTEM",
            "capex": 185_000_000, "opex_rate": 0.06, "opex_per_ton": 85_000,
            "capacity_kg_hr": 250, "is_core": True,
            "format_compat": ["SSS","BIB","STICKPACK"], "line_compat": ["B","G","STICKPACK"], "role": "Inspeksi", "url": "",
        },
    }

# ── Overhead breakdown (dari CAPEX GENERAL, fallback 18%) ────────────────────
if _fp_capex_gen:
    OVERHEAD = {k: v["value"] for k, v in _fp_capex_gen.items()
                if v["basis"] == "PERCENT_MACHINE"}
    COMMISSIONING_FIXED = sum(v["value"] for v in _fp_capex_gen.values()
                              if v["basis"] == "FIXED")
else:
    OVERHEAD = {"Instalasi": 0.08, "Elektrikal": 0.05, "Utilitas": 0.03, "Sparepart Awal": 0.02}
    COMMISSIONING_FIXED = 50_000_000

OVERHEAD_TOTAL = sum(OVERHEAD.values())

# ── OPEX manpower (dari OPEX GENERAL) ────────────────────────────────────────
if _fp_opex_gen:
    ANNUAL_OPERATOR = int(_fp_opex_gen.get("FEE_OPERATOR_NEWLINE", {}).get("value", 83_200_000))
    ANNUAL_QC       = int(_fp_opex_gen.get("FEE_QC_NEWLINE",       {}).get("value", 83_200_000))
else:
    ANNUAL_OPERATOR = 83_200_000
    ANNUAL_QC       = 83_200_000

ANNUAL_MAINTENANCE_FBMI = 240_000_000  # Pak Ardi: Rp 20M/bulan (konfirmasi langsung)

# ── Default financial parameters ──────────────────────────────────────────────
DEFAULT_PARAMS = {
    "discount_rate":            float(_fp_financial.get("DISCOUNT RATE",            0.12)),
    "project_lifetime_year":    int(_fp_financial.get("PROJECT LIFETIME YEAR",       5)),
    "payback_threshold_year":   int(_fp_financial.get("PAYBACK THRESHOLD YEAR",      3)),
    "minimum_npv":              float(_fp_financial.get("MINIMUM NPV",               0)),
    "minimum_irr":              float(_fp_financial.get("MINIMUM IRR",               0.15)),
    "minimum_roi":              float(_fp_financial.get("MINIMUM ROI",               0.25)),
    "realization_factor":       float(_fp_financial.get("REALIZATION FACTOR",        0.75)),
    "internal_value_per_ton":   float(_fp_financial.get("INTERNAL VALUE PER TON",    2_100_000)),
    "maklon_cost_per_ton":      float(_fp_financial.get("MAKLON COST PER TON",       6_500_000)),
    "internal_cost_per_ton":    float(_fp_financial.get("INTERNAL COST PER TON",     5_000_000)),
    "tax_rate":                 0.25,
    "useful_life_year":         int(_fp_financial.get("PROJECT LIFETIME YEAR",       5)),
    "maintenance_annual":       ANNUAL_MAINTENANCE_FBMI,
}


def _machine_capex(machines_qty: list, machines_dict: dict = None) -> dict:
    """machines_qty: list of (machine_key, qty)"""
    md = machines_dict or MACHINES
    total_machine = 0
    breakdown     = {}
    annual_maint  = 0
    for key, qty in machines_qty:
        m    = md.get(key, {})
        cost = m.get("capex", 0) * qty
        total_machine += cost
        maint = m.get("capex", 0) * m.get("opex_rate", 0.05) * qty
        annual_maint += maint
        label = m.get("full_name", key) + (f" ×{qty}" if qty > 1 else "")
        breakdown[label] = cost
    overhead = total_machine * OVERHEAD_TOTAL
    for k, v in OVERHEAD.items():
        breakdown[k] = int(total_machine * v)
    breakdown["Komisioning & Training"] = int(COMMISSIONING_FIXED)
    return {
        "machine":              total_machine,
        "overhead":             overhead,
        "commissioning":        COMMISSIONING_FIXED,
        "breakdown":            breakdown,
        "annual_maintenance":   annual_maint,
    }


def capex_from_catalog(cat: dict, pkg_key: str) -> dict:
    """Hitung CAPEX dari katalog mesin (investment_catalog.json)."""
    machines  = cat.get("machines", MACHINES)
    pkg       = cat.get("packages", {}).get(pkg_key, {})
    components = pkg.get("components", [])
    machines_qty = [(c["key"], c.get("qty", 1)) for c in components]
    r = _machine_capex(machines_qty, machines)
    maint_annual = pkg.get("maintenance_annual", ANNUAL_MAINTENANCE_FBMI)
    r["total"]              = int(r["machine"] + r["overhead"] + r["commissioning"])
    r["annual_opex_maint"]  = maint_annual
    r["annual_opex_staff"]  = ANNUAL_OPERATOR + ANNUAL_QC
    r["annual_opex_total"]  = maint_annual + ANNUAL_OPERATOR + ANNUAL_QC
    r["pkg_name"]           = pkg.get("name", pkg_key)
    r["machine_list"]       = machines_qty
    return r


def capex_multiline(qty_lines: int = 1) -> dict:
    """CAPEX konversi single → multiline (kompatibel dengan kode lama)."""
    # Ambil mesin yang relevan dari catalog
    filler_keys  = [k for k, m in MACHINES.items()
                    if "SSS" in m.get("format_compat", []) and m.get("is_core", False)
                    and "filler" in m.get("full_name","").lower()]
    if not filler_keys:
        filler_keys = [k for k, m in MACHINES.items() if "SSS" in m.get("format_compat",[])]
    # Default: SHIPUTEC multi-lane + micro auger + conveyor
    machines = []
    for k in filler_keys[:1]:
        machines.append((k, qty_lines))
    check_keys = [k for k, m in MACHINES.items() if "CHECKWEIGHER" in m.get("full_name","").upper()]
    xray_keys  = [k for k, m in MACHINES.items() if "XRAY" in m.get("full_name","").upper() or "X-RAY" in m.get("full_name","").upper()]
    if check_keys: machines.append((check_keys[0], qty_lines))
    if xray_keys:  machines.append((xray_keys[0],  qty_lines))

    r = _machine_capex(machines)
    r["total"]             = int(r["machine"] + r["overhead"] + r["commissioning"])
    r["annual_opex_maint"] = ANNUAL_MAINTENANCE_FBMI
    r["annual_opex_staff"] = 0
    r["annual_opex_total"] = ANNUAL_MAINTENANCE_FBMI
    r["machine_list"]      = machines
    return r


def capex_new_line() -> dict:
    """CAPEX penambahan 1 lini baru single-lane SSS+BIB."""
    filler_keys = [k for k, m in MACHINES.items()
                   if "SSS" in m.get("format_compat",[]) and "BIB" in m.get("format_compat",[])
                   and m.get("is_core", False)]
    machines = []
    if filler_keys:
        machines.append((filler_keys[0], 1))
    check_keys = [k for k, m in MACHINES.items() if "CHECKWEIGHER" in m.get("full_name","").upper()]
    xray_keys  = [k for k, m in MACHINES.items() if "XRAY" in m.get("full_name","").upper() or "X-RAY" in m.get("full_name","").upper()]
    if check_keys: machines.append((check_keys[0], 1))
    if xray_keys:  machines.append((xray_keys[0],  1))
    r = _machine_capex(machines)
    r["total"]             = int(r["machine"] + r["overhead"] + r["commissioning"])
    r["annual_opex_maint"] = ANNUAL_MAINTENANCE_FBMI
    r["annual_opex_staff"] = ANNUAL_OPERATOR + ANNUAL_QC
    r["annual_opex_total"] = ANNUAL_MAINTENANCE_FBMI + ANNUAL_OPERATOR + ANNUAL_QC
    r["machine_list"]      = machines
    return r


def capex_stickpack_line() -> dict:
    """CAPEX lini stickpack baru."""
    stick_keys = [k for k, m in MACHINES.items()
                  if "STICKPACK" in m.get("format_compat",[]) and m.get("is_core", False)
                  and "filler" in m.get("full_name","").lower()]
    machines = []
    if stick_keys:
        machines.append((stick_keys[0], 1))
    check_keys = [k for k, m in MACHINES.items() if "CHECKWEIGHER" in m.get("full_name","").upper()]
    xray_keys  = [k for k, m in MACHINES.items() if "XRAY" in m.get("full_name","").upper() or "X-RAY" in m.get("full_name","").upper()]
    if check_keys: machines.append((check_keys[0], 1))
    if xray_keys:  machines.append((xray_keys[0],  1))
    r = _machine_capex(machines)
    r["total"]             = int(r["machine"] + r["overhead"] + r["commissioning"])
    r["annual_opex_maint"] = ANNUAL_MAINTENANCE_FBMI
    r["annual_opex_staff"] = ANNUAL_OPERATOR + ANNUAL_QC
    r["annual_opex_total"] = ANNUAL_MAINTENANCE_FBMI + ANNUAL_OPERATOR + ANNUAL_QC
    r["machine_list"]      = machines
    return r


CAPEX_FN = {
    "multiline_G":  lambda: capex_multiline(1),
    "multiline_B":  lambda: capex_multiline(1),
    "multiline_BG": lambda: capex_multiline(2),
    "new_line":     capex_new_line,
    "stickpack_line": capex_stickpack_line,
    "multiline_BG_new": lambda: {
        "total": capex_multiline(2)["total"] + capex_new_line()["total"],
        "annual_opex_total": capex_multiline(2)["annual_opex_total"] + capex_new_line()["annual_opex_total"],
        "breakdown": {**capex_multiline(2)["breakdown"],
                      **{f"[Lini Baru] {k}": v for k,v in capex_new_line()["breakdown"].items()}},
        "machine_list": capex_multiline(2)["machine_list"] + capex_new_line()["machine_list"],
    },
}


def compute_financial(total_capex: float,
                      annual_additional_ton: float,
                      params: dict,
                      annual_opex_extra: float = 0.0) -> dict:
    """
    DCF valuation sesuai model Pak Ardi (Fonterra/Lactalis Group Valuation Model).

    Alur perhitungan (dari Valuation sheet):
        EBIT       = Benefit - OPEX - Depreciation
        Cash Tax   = max(EBIT, 0) × tax_rate   [tax shield via dep]
        FCF/tahun  = EBIT + Depreciation - Cash Tax - Maintenance Capex
                   = Benefit - OPEX - Cash Tax - Maintenance Capex

    Benefit = ton × value_per_ton × realization_factor
    Untuk maklon → internalisasi: value = maklon_cost - internal_cost (per ton)
    """
    r        = float(params.get("discount_rate",          DEFAULT_PARAMS["discount_rate"]))
    N        = int(params.get("project_lifetime_year",    DEFAULT_PARAMS["project_lifetime_year"]))
    rf       = float(params.get("realization_factor",     DEFAULT_PARAMS["realization_factor"]))
    vpt      = float(params.get("internal_value_per_ton", DEFAULT_PARAMS["internal_value_per_ton"]))
    tax_rate = float(params.get("tax_rate",               DEFAULT_PARAMS["tax_rate"]))
    ul       = max(int(params.get("useful_life_year",     DEFAULT_PARAMS["useful_life_year"])), 1)
    maint    = float(params.get("maintenance_annual",     DEFAULT_PARAMS["maintenance_annual"]))

    # Benefit override (untuk kasus maklon)
    if params.get("_benefit_override", 0) > 0:
        annual_benefit = float(params["_benefit_override"])
    else:
        annual_benefit = annual_additional_ton * vpt * rf

    # Straight-line depreciation (tax shield, add back ke FCF)
    depreciation = total_capex / ul

    # Income statement
    ebit      = annual_benefit - annual_opex_extra - depreciation
    cash_tax  = max(ebit, 0.0) * tax_rate

    # FCF per tahun = EBIT + Dep - Cash Tax - Maintenance Capex (per model Pak Ardi)
    annual_fcf = annual_benefit - annual_opex_extra - cash_tax - maint

    # Cash flow series: t=0 outflow CAPEX, t=1..N inflow FCF
    cash_flows = [-total_capex] + [annual_fcf] * N

    # NPV dengan discount rate = WACC
    npv = float(npf.npv(r, cash_flows))

    # IRR
    try:
        irr_val = npf.irr(cash_flows)
        irr_pct = float(irr_val * 100) if (irr_val is not None and not np.isnan(irr_val)) else None
    except Exception:
        irr_pct = None

    # ROI = FCF tahunan / CAPEX (return tahunan atas modal)
    roi_pct = (annual_fcf / total_capex * 100) if total_capex > 0 else 0

    # Payback period (undiscounted, sesuai output Pak Ardi)
    payback = (total_capex / annual_fcf) if annual_fcf > 0 else None

    # Kelayakan
    min_irr  = params.get("minimum_irr",            DEFAULT_PARAMS["minimum_irr"])
    min_roi  = params.get("minimum_roi",             DEFAULT_PARAMS["minimum_roi"])
    min_npv  = params.get("minimum_npv",             DEFAULT_PARAMS["minimum_npv"])
    pb_thr   = params.get("payback_threshold_year",  DEFAULT_PARAMS["payback_threshold_year"])

    flags = {
        f"NPV ≥ {fmt_rp(min_npv)}":      npv >= min_npv,
        f"IRR ≥ {min_irr*100:.0f}%":     (irr_pct is not None and irr_pct/100 >= min_irr),
        f"ROI ≥ {min_roi*100:.0f}%":     roi_pct/100 >= min_roi,
        f"Payback ≤ {pb_thr} thn":       (payback is not None and payback <= pb_thr),
    }
    feasible = flags[f"NPV ≥ {fmt_rp(min_npv)}"] and flags[f"IRR ≥ {min_irr*100:.0f}%"]

    return {
        "npv":                  npv,
        "irr_pct":              irr_pct,
        "roi_pct":              roi_pct,
        "payback_year":         payback,
        "annual_benefit":       annual_benefit,
        "annual_fcf":           annual_fcf,
        "annual_opex_extra":    annual_opex_extra,
        "depreciation":         depreciation,
        "ebit":                 ebit,
        "cash_tax":             cash_tax,
        "annual_maint_capex":   maint,
        "cash_flows":           cash_flows,
        "flags":                flags,
        "feasible":             feasible,
        "annual_additional_ton": annual_additional_ton,
        "realization_factor":   rf,
        "N":                    N,
        "discount_rate":        r,
    }


def compute_financial_sensitivity(total_capex, annual_ton, params,
                                  annual_opex_extra=0, scenarios=None):
    """
    Hitung kelayakan finansial untuk beberapa skenario sensitivitas.
    Scenarios: list of (label, ton_multiplier, discount_rate_delta)
    """
    if scenarios is None:
        scenarios = [
            ("Optimis (+20%)",   1.20, -0.01),
            ("Base Case",        1.00,  0.00),
            ("Konservatif (-20%)",0.80, +0.01),
        ]
    results = []
    for label, ton_mult, dr_delta in scenarios:
        p = dict(params)
        p["discount_rate"] = float(p.get("discount_rate", DEFAULT_PARAMS["discount_rate"])) + dr_delta
        res = compute_financial(total_capex, annual_ton * ton_mult, p, annual_opex_extra)
        results.append({"Skenario": label, **res})
    return results


def monte_carlo_npv(total_capex, monthly_mean, monthly_std, params, n_sim=5000):
    np.random.seed(42)
    r   = float(params.get("discount_rate",          DEFAULT_PARAMS["discount_rate"]))
    N   = int(params.get("project_lifetime_year",    DEFAULT_PARAMS["project_lifetime_year"]))
    rf  = float(params.get("realization_factor",     DEFAULT_PARAMS["realization_factor"]))
    vpt = float(params.get("internal_value_per_ton", DEFAULT_PARAMS["internal_value_per_ton"]))
    maint = float(params.get("maintenance_annual",   DEFAULT_PARAMS["maintenance_annual"]))
    samples = np.clip(np.random.normal(monthly_mean, max(monthly_std, 0.1), n_sim), 0, None)
    npvs = np.array([npf.npv(r, [-total_capex] + [s*12*vpt*rf - maint]*N) for s in samples])
    return {
        "npv_mean":       npvs.mean(),
        "npv_p5":         np.percentile(npvs, 5),
        "npv_p25":        np.percentile(npvs, 25),
        "npv_p75":        np.percentile(npvs, 75),
        "npv_p95":        np.percentile(npvs, 95),
        "prob_positive":  (npvs > 0).mean() * 100,
        "samples":        npvs,
    }
